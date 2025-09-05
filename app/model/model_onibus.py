"""
Arquivo: model_onibus.py
Responsável pela lógica de dados e manipulação das reservas do ônibus.
"""

from openpyxl import load_workbook


class Onibus:
    def __init__(self, capacidade, caminho_arquivo):
        """
        Inicializa o objeto Onibus com a capacidade de lugares e o caminho do arquivo de dados.
        """
        self.capacidade = capacidade
        self.lugares = [0] * capacidade
        self.caminho_arquivo = caminho_arquivo

    def reservar_lugar(self, num_lugar, nome, cpf, dia):
        """
        Realiza a reserva de um lugar, se disponível, e salva no arquivo.
        Retorna mensagem de sucesso ou erro.
        """
        if num_lugar < 1 or num_lugar > self.capacidade:
            return "Lugar inválido"
        if self.lugares[num_lugar - 1] == 0:
            self.lugares[num_lugar - 1] = 1
            self.salvar_reserva(num_lugar, nome, cpf, dia)
            return f"Lugar {num_lugar} reservado com sucesso."
        else:
            return f"Lugar {num_lugar} indisponível."

    def cancelar_reserva(self, num_lugar, dia):
        """
        Cancela a reserva de um lugar, se estiver reservado, e remove do arquivo.
        Retorna mensagem de sucesso ou erro.
        """
        if num_lugar < 1 or num_lugar > self.capacidade:
            return "Lugar inválido"
        if self.lugares[num_lugar - 1] == 1:
            self.lugares[num_lugar - 1] = 0
            self.excluir_reserva(num_lugar, dia)
            return f"Lugar {num_lugar} reserva cancelada com sucesso."
        else:
            return f"Lugar {num_lugar} não está reservado."

    def salvar_reserva(self, num_lugar, nome, cpf, dia):
        """
        Salva uma nova reserva no arquivo Excel.
        """
        try:
            workbook = load_workbook(self.caminho_arquivo)
            sheet = workbook['Reservas']
        except FileNotFoundError:
            return
        linha = [num_lugar, nome, cpf, dia]
        sheet.append(linha)
        workbook.save(self.caminho_arquivo)

    def excluir_reserva(self, num_lugar, dia):
        """
        Remove uma reserva do arquivo Excel com base no número do lugar e dia.
        """
        try:
            workbook = load_workbook(self.caminho_arquivo)
            sheet = workbook['Reservas']
        except FileNotFoundError:
            return
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == num_lugar and row[3].value == dia:
                sheet.delete_rows(row[0].row)
                workbook.save(self.caminho_arquivo)
                break

    def carregar_reservas(self, dia):
        """
        Carrega as reservas do arquivo Excel para o dia informado e atualiza o status dos lugares.
        """
        try:
            workbook = load_workbook(self.caminho_arquivo)
            sheet = workbook['Reservas']
        except FileNotFoundError:
            return
        self.lugares = [0] * self.capacidade
        for i, linha in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:
                continue
            num_lugar = linha[0]
            if num_lugar and 1 <= num_lugar <= self.capacidade:
                if linha[3] == dia:
                    self.lugares[num_lugar - 1] = 1

    def gerar_mapa(self):
        """
        Gera uma representação visual do mapa de assentos do ônibus, indicando quais estão reservados.
        """
        mapa = ""
        for i in range(0, self.capacidade, 2):
            lugar_esquerda = i + 1
            lugar_direita = i + 2
            status_esquerda = "X" if self.lugares[i] == 1 else " "
            status_direita = "X" if lugar_direita <= self.capacidade and self.lugares[i + 1] == 1 else " "
            mapa += f"Lugar {lugar_esquerda}: [{status_esquerda}]    Lugar {lugar_direita}: [{status_direita}]\n"
        return mapa
